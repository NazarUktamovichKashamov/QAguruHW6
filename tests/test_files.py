import csv
from io import TextIOWrapper

from openpyxl import load_workbook
from zipfile import ZipFile
from pypdf import PdfReader
import pandas as pd

from set_path import ZIP_FILE


def test_check_pdf_in_archive():
    file_name = 'sample2.pdf'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            reader = PdfReader(file)

            assert len(reader.pages) == 1
            assert "PDF Form Example" in reader.pages[0].extract_text()


def test_check_xlsx_in_archive():
    file_name = 'sample1.xlsx'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            workbook = load_workbook(file)
            title_list = workbook.sheetnames
            sheet = workbook.active

            assert sheet.cell(row=2, column=2).value == 456


def test_check_csv_in_archive_csv():
    file_name = 'sample3.csv'
    with ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open(file_name) as file:
            csv_file = file.read().decode(encoding='utf-8')
            content = list(csv.reader(csv_file.splitlines(), delimiter=';'))

            assert content[0][0] == 'Game Number, "Game Length"'

